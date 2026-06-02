from pathlib import Path
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from copy import copy

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS")
OUTPUT_DIR = BASE / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "District_Employee_Transfer_Database_Template.xlsx"


COLORS = {
    "navy": "1F4E78",
    "blue": "5B9BD5",
    "light_blue": "DDEBF7",
    "green": "C6EFCE",
    "yellow": "FFEB9C",
    "red": "F4CCCC",
    "gray": "E7E6E6",
    "dark_gray": "595959",
    "white": "FFFFFF",
    "black": "000000",
    "orange": "FCE4D6",
}


def style_title(cell, size=18):
    cell.font = Font(bold=True, size=size, color=COLORS["white"])
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section(cell, fill=COLORS["light_blue"]):
    cell.font = Font(bold=True, color=COLORS["black"])
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(row_cells):
    for cell in row_cells:
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="A6A6A6"))


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name, ref):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_workbook():
    wb = Workbook()
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    ws_master = wb.create_sheet("Employee Master")
    ws_history = wb.create_sheet("Posting History")
    ws_reports = wb.create_sheet("Reports")
    ws_lists = wb.create_sheet("Lists")
    ws_guide = wb.create_sheet("Guide")

    build_lists(ws_lists)
    add_named_ranges(wb)
    build_master(ws_master)
    build_history(ws_history)
    build_dashboard(ws_dashboard)
    build_reports(ws_reports)
    build_guide(ws_guide)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

    ws_dashboard.freeze_panes = "A4"
    ws_history.freeze_panes = "A2"
    ws_master.freeze_panes = "A2"
    ws_reports.freeze_panes = "A4"

    return wb


def add_named_ranges(wb):
    named_ranges = {
        "RankList": "Lists!$A$4:$A$20",
        "DistrictList": "Lists!$B$4:$B$20",
        "SubDivisionList": "Lists!$C$4:$C$20",
        "StationList": "Lists!$D$4:$D$30",
        "KGIDList": "'Employee Master'!$A$2:$A$501",
    }
    for name, ref in named_ranges.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def build_lists(ws):
    ws["A1"] = "Reference Lists"
    style_title(ws["A1"])
    ws.merge_cells("A1:F1")
    data = {
        "A": ("Ranks / Designations", ["PC", "HC", "ASI", "PSI", "CPI", "PI"]),
        "B": ("Districts", ["Kolar", "Bengaluru Rural", "Chikkaballapura"]),
        "C": ("Sub-Divisions", ["KGF", "Bangarpet", "Malur", "Kolar"]),
        "D": ("Police Stations / Units", ["Robertsonpet PS", "Bethamangala PS", "Bangarpet PS", "Malur PS", "Kolar Town PS", "District Armed Reserve"]),
        "E": ("Priority", ["Low", "Medium", "High"]),
        "F": ("Document Types", ["Appointment Order", "Transfer Order", "Promotion Order", "Service Register Copy"]),
    }
    for col, (header, values) in data.items():
        ws[f"{col}3"] = header
        style_section(ws[f"{col}3"])
        for i, value in enumerate(values, start=4):
            ws[f"{col}{i}"] = value
    set_widths(ws, {"A": 24, "B": 22, "C": 22, "D": 30, "E": 16, "F": 26})


def build_master(ws):
    headers = [
        "KGID", "Employee Name", "Rank", "Designation", "DOB", "Appointment Date",
        "Mobile", "Home District", "Retirement Date", "Current District",
        "Current Sub-Division", "Current Unit", "Present Posting Date",
        "Total Service Years", "Current Station Years", "Current Sub-Division Years",
        "District Service Years", "Transfer Eligible", "Priority", "Remarks"
    ]
    ws.append(headers)
    style_header(ws[1])

    sample = [
        ["12345", "Ravi Kumar", "HC", "Head Constable", date(1980, 2, 21), date(2005, 6, 15), "9876543210", "Kolar", None, "Kolar", "KGF", "Robertsonpet PS", date(2023, 8, 15), None, None, None, None, None, None, ""],
        ["12346", "Anil Gowda", "PC", "Police Constable", date(1988, 9, 8), date(2012, 4, 20), "9876500001", "Chikkaballapura", None, "Kolar", "Bangarpet", "Bangarpet PS", date(2019, 7, 1), None, None, None, None, None, None, ""],
        ["12347", "Suresh Naik", "ASI", "Assistant Sub Inspector", date(1975, 11, 12), date(1999, 2, 5), "9876500002", "Bengaluru Rural", None, "Kolar", "Malur", "Malur PS", date(2016, 1, 10), None, None, None, None, None, None, ""],
        ["12348", "Meena Patil", "PSI", "Police Sub Inspector", date(1983, 5, 30), date(2008, 10, 1), "9876500003", "Kolar", None, "Kolar", "Kolar", "Kolar Town PS", date(2022, 3, 17), None, None, None, None, None, None, ""],
        ["12349", "Nagaraj Rao", "HC", "Head Constable", date(1970, 12, 15), date(1996, 8, 12), "9876500004", "Kolar", None, "Kolar", "KGF", "Bethamangala PS", date(2018, 5, 25), None, None, None, None, None, None, ""],
    ]
    for row in sample:
        ws.append(row)

    for row in range(2, 502):
        ws[f"I{row}"] = f'=IF(E{row}="","",EOMONTH(EDATE(E{row},60*12),0))'
        ws[f"N{row}"] = f'=IF(F{row}="","",ROUND(YEARFRAC(F{row},TODAY()),1))'
        ws[f"O{row}"] = f'=IF(M{row}="","",ROUND(YEARFRAC(M{row},TODAY()),1))'
        ws[f"P{row}"] = f'=IF(A{row}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$1000=A{row})*(\'Posting History\'!$E$2:$E$1000=K{row})*(\'Posting History\'!$B$2:$B$1000<>"")*(IF(\'Posting History\'!$C$2:$C$1000="",TODAY(),\'Posting History\'!$C$2:$C$1000)-\'Posting History\'!$B$2:$B$1000))/365.25,1))'
        ws[f"Q{row}"] = f'=IF(A{row}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$1000=A{row})*(\'Posting History\'!$F$2:$F$1000=J{row})*(\'Posting History\'!$B$2:$B$1000<>"")*(IF(\'Posting History\'!$C$2:$C$1000="",TODAY(),\'Posting History\'!$C$2:$C$1000)-\'Posting History\'!$B$2:$B$1000))/365.25,1))'
        ws[f"R{row}"] = f'=IF(O{row}="","",IF(O{row}>=3,"Yes","No"))'
        ws[f"S{row}"] = f'=IF(O{row}="","",IF(O{row}>=5,"High",IF(O{row}>=3,"Medium","Low")))'

    for col in ["E", "F", "I", "M"]:
        for cell in ws[col][1:501]:
            cell.number_format = "dd-mm-yyyy"
    for col in ["N", "O", "P", "Q"]:
        for cell in ws[col][1:501]:
            cell.number_format = "0.0"

    add_table(ws, "tblEmployeeMaster", "A1:T501")
    set_widths(ws, {
        "A": 12, "B": 22, "C": 12, "D": 24, "E": 14, "F": 18, "G": 14, "H": 18,
        "I": 16, "J": 18, "K": 20, "L": 24, "M": 18, "N": 16, "O": 18, "P": 22,
        "Q": 18, "R": 16, "S": 12, "T": 28,
    })
    add_master_validations(ws)
    add_priority_coloring(ws, "S2:S501")


def build_history(ws):
    headers = [
        "KGID", "From Date", "To Date", "Police Station / Unit", "Sub-Division",
        "District", "Rank Held", "Order Number", "Duration Years", "Overlap Check", "Missing Date Check", "Notes"
    ]
    ws.append(headers)
    style_header(ws[1])
    rows = [
        ["12345", date(2005, 6, 15), date(2008, 8, 10), "PS A", "KGF", "Kolar", "PC", "TR-101", None, None, None, ""],
        ["12345", date(2008, 8, 11), date(2013, 4, 1), "PS B", "Bangarpet", "Kolar", "PC", "TR-204", None, None, None, ""],
        ["12345", date(2013, 4, 2), date(2020, 9, 18), "PS C", "Malur", "Kolar", "HC", "PR-008", None, None, None, ""],
        ["12345", date(2020, 9, 19), date(2023, 8, 14), "Bethamangala PS", "KGF", "Kolar", "HC", "TR-330", None, None, None, ""],
        ["12345", date(2023, 8, 15), None, "Robertsonpet PS", "KGF", "Kolar", "HC", "TR-415", None, None, None, ""],
        ["12346", date(2012, 4, 20), date(2019, 6, 30), "District Armed Reserve", "Kolar", "Kolar", "PC", "", None, None, None, ""],
        ["12346", date(2019, 7, 1), None, "Bangarpet PS", "Bangarpet", "Kolar", "PC", "", None, None, None, ""],
        ["12347", date(1999, 2, 5), date(2016, 1, 9), "Kolar Town PS", "Kolar", "Kolar", "HC", "", None, None, None, ""],
        ["12347", date(2016, 1, 10), None, "Malur PS", "Malur", "Kolar", "ASI", "", None, None, None, ""],
        ["12348", date(2008, 10, 1), date(2022, 3, 16), "Robertsonpet PS", "KGF", "Kolar", "PSI", "", None, None, None, ""],
        ["12348", date(2022, 3, 17), None, "Kolar Town PS", "Kolar", "Kolar", "PSI", "", None, None, None, ""],
        ["12349", date(1996, 8, 12), date(2018, 5, 24), "Bangarpet PS", "Bangarpet", "Kolar", "PC", "", None, None, None, ""],
        ["12349", date(2018, 5, 25), None, "Bethamangala PS", "KGF", "Kolar", "HC", "", None, None, None, ""],
    ]
    for row in rows:
        ws.append(row)
    for row in range(2, 1001):
        ws[f"I{row}"] = f'=IF(B{row}="","",ROUND((IF(C{row}="",TODAY(),C{row})-B{row})/365.25,1))'
        ws[f"J{row}"] = f'=IF(A{row}="","",IF(COUNTIFS($A:$A,A{row},$B:$B,"<="&IF(C{row}="",TODAY(),C{row}),$C:$C,">="&B{row})>1,"Review","OK"))'
        ws[f"K{row}"] = f'=IF(A{row}="","",IF(OR(B{row}="",D{row}="",E{row}="",F{row}="",G{row}=""),"Missing","OK"))'
    for col in ["B", "C"]:
        for cell in ws[col][1:1000]:
            cell.number_format = "dd-mm-yyyy"
    for cell in ws["I"][1:1000]:
        cell.number_format = "0.0"
    add_table(ws, "tblPostingHistory", "A1:L1000")
    set_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 24, "E": 18, "F": 16, "G": 14, "H": 18, "I": 14, "J": 15, "K": 18, "L": 28})
    add_history_validations(ws)
    add_status_coloring(ws)


def add_dropdowns(ws, definitions):
    for source, ranges in definitions:
        dv = DataValidation(type="list", formula1=source, allow_blank=True)
        ws.add_data_validation(dv)
        for rng in ranges.split():
            dv.add(rng)


def add_master_validations(ws):
    add_dropdowns(ws, [
        ("RankList", "C2:C1000"),
        ("DistrictList", "H2:H1000 J2:J1000"),
        ("SubDivisionList", "K2:K1000"),
        ("StationList", "L2:L1000"),
    ])


def add_history_validations(ws):
    add_dropdowns(ws, [
        ("RankList", "G2:G1000"),
        ("DistrictList", "F2:F1000"),
        ("SubDivisionList", "E2:E1000"),
        ("StationList", "D2:D1000"),
    ])


def add_priority_coloring(ws, rng):
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'S2="Low"'], fill=PatternFill("solid", fgColor=COLORS["green"])))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'S2="Medium"'], fill=PatternFill("solid", fgColor=COLORS["yellow"])))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'S2="High"'], fill=PatternFill("solid", fgColor=COLORS["red"])))


def add_status_coloring(ws):
    ws.conditional_formatting.add("J2:J1000", FormulaRule(formula=['J2="Review"'], fill=PatternFill("solid", fgColor=COLORS["red"])))
    ws.conditional_formatting.add("K2:K1000", FormulaRule(formula=['K2="Missing"'], fill=PatternFill("solid", fgColor=COLORS["yellow"])))


def build_dashboard(ws):
    ws["A1"] = "District Employee Transfer Dashboard"
    ws.merge_cells("A1:J1")
    style_title(ws["A1"], 20)
    ws["A2"] = "Search by KGID, name, mobile, date of appointment, designation, or present station. Select a KGID in C12 to display the full employee profile below."
    ws.merge_cells("A2:J2")
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["gray"])
    ws["A2"].alignment = Alignment(wrap_text=True)

    search_rows = [
        ("A4", "Search Section", None),
        ("A5", "KGID", "C5"),
        ("A6", "Employee Name", "C6"),
        ("A7", "Mobile Number", "C7"),
        ("A8", "Date of Appointment", "C8"),
        ("A9", "Designation", "C9"),
        ("A10", "Present Station", "C10"),
        ("A12", "Selected KGID", "C12"),
    ]
    for addr, label, input_cell in search_rows:
        ws[addr] = label
        style_section(ws[addr], COLORS["light_blue"] if addr == "A4" else COLORS["gray"])
        if input_cell:
            ws[input_cell].fill = PatternFill("solid", fgColor=COLORS["white"])
            ws[input_cell].border = Border(bottom=Side(style="thin", color="808080"))

    ws["C8"].number_format = "dd-mm-yyyy"
    for cell, source in [("C9", "RankList"), ("C10", "StationList"), ("C12", "KGIDList")]:
        dv = DataValidation(type="list", formula1=source, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws[cell])

    ws["E4"] = "Search Results"
    style_section(ws["E4"])
    result_headers = ["KGID", "Name", "Rank", "Mobile", "DOA", "Current Unit", "Current Since", "Station Years", "Eligible"]
    for idx, header in enumerate(result_headers, start=5):
        ws.cell(5, idx).value = header
    style_header(ws[5][4:13])
    for row in range(6, 21):
        src = row - 4
        ws[f"E{row}"] = f'=IF(OR($C$5<>"",$C$6<>"",$C$7<>"",$C$8<>"",$C$9<>"",$C$10<>""),IF(AND(IF($C$5="",TRUE,ISNUMBER(SEARCH($C$5,\'Employee Master\'!$A{src}))),IF($C$6="",TRUE,ISNUMBER(SEARCH($C$6,\'Employee Master\'!$B{src}))),IF($C$7="",TRUE,ISNUMBER(SEARCH($C$7,\'Employee Master\'!$G{src}))),IF($C$8="",TRUE,\'Employee Master\'!$F{src}=$C$8),IF($C$9="",TRUE,ISNUMBER(SEARCH($C$9,\'Employee Master\'!$D{src}))),IF($C$10="",TRUE,ISNUMBER(SEARCH($C$10,\'Employee Master\'!$L{src})))),\'Employee Master\'!$A{src},""),"")'
        ws[f"F{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$B:$B,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"G{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$C:$C,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"H{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$G:$G,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"I{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$F:$F,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"J{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$L:$L,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"K{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$M:$M,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"L{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$O:$O,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'
        ws[f"M{row}"] = f'=IF($E{row}="","",INDEX(\'Employee Master\'!$R:$R,MATCH($E{row},\'Employee Master\'!$A:$A,0)))'

    profile_sections = [
        (15, "Basic Information"),
        (24, "Current Posting"),
        (33, "Complete Posting History"),
        (48, "Transfer Information"),
        (58, "Documents"),
    ]
    for row, title in profile_sections:
        ws[f"A{row}"] = f"+ / - {title}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        style_section(ws[f"A{row}"], COLORS["blue"])
        ws[f"A{row}"].font = Font(bold=True, color=COLORS["white"])

    detail_formulas = [
        (16, "KGID", '=C12'),
        (17, "Name", '=IFERROR(INDEX(\'Employee Master\'!$B:$B,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (18, "Rank", '=IFERROR(INDEX(\'Employee Master\'!$C:$C,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (19, "Mobile", '=IFERROR(INDEX(\'Employee Master\'!$G:$G,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (20, "DOB", '=IFERROR(INDEX(\'Employee Master\'!$E:$E,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (21, "DOA", '=IFERROR(INDEX(\'Employee Master\'!$F:$F,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (22, "Retirement Date", '=IFERROR(INDEX(\'Employee Master\'!$I:$I,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (25, "District", '=IFERROR(INDEX(\'Employee Master\'!$J:$J,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (26, "Sub Division", '=IFERROR(INDEX(\'Employee Master\'!$K:$K,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (27, "Police Station / Unit", '=IFERROR(INDEX(\'Employee Master\'!$L:$L,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (28, "Joined On", '=IFERROR(INDEX(\'Employee Master\'!$M:$M,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (29, "Service in Present Station", '=IFERROR(INDEX(\'Employee Master\'!$O:$O,MATCH($C$12,\'Employee Master\'!$A:$A,0))&" Years","")'),
        (49, "Total Service", '=IFERROR(INDEX(\'Employee Master\'!$N:$N,MATCH($C$12,\'Employee Master\'!$A:$A,0))&" Years","")'),
        (50, "Stations Served", '=IF($C$12="","",COUNTIF(\'Posting History\'!$A:$A,$C$12))'),
        (51, "Longest Posting", '=IF($C$12="","",MAXIFS(\'Posting History\'!$I:$I,\'Posting History\'!$A:$A,$C$12)&" Years")'),
        (52, "Eligible for Transfer", '=IFERROR(INDEX(\'Employee Master\'!$R:$R,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
        (53, "Priority", '=IFERROR(INDEX(\'Employee Master\'!$S:$S,MATCH($C$12,\'Employee Master\'!$A:$A,0)),"")'),
    ]
    for row, label, formula in detail_formulas:
        ws[f"A{row}"] = label
        ws[f"C{row}"] = formula
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"].fill = PatternFill("solid", fgColor=COLORS["white"])

    history_headers = ["Sl", "From", "To", "Station", "Sub Division", "District", "Rank", "Duration"]
    for idx, header in enumerate(history_headers, start=1):
        ws.cell(34, idx).value = header
    style_header(ws[34][:8])
    for row in range(35, 46):
        src = row - 33
        ws[f"A{row}"] = f'=IF(\'Posting History\'!$A{src}=$C$12,COUNTIF(\'Posting History\'!$A$2:$A{src},$C$12),"")'
        for offset, hist_col in enumerate(["B", "C", "D", "E", "F", "G", "I"], start=2):
            letter = get_column_letter(offset)
            ws[f"{letter}{row}"] = f'=IF($A{row}="","",\'Posting History\'!${hist_col}{src})'

    docs = ["Appointment Order", "Transfer Orders", "Promotion Orders", "Service Register Copy"]
    for idx, doc in enumerate(docs, start=59):
        ws[f"A{idx}"] = doc
        ws[f"C{idx}"] = "File link / reference"

    for start, end in [(16, 22), (25, 29), (34, 45), (49, 53), (59, 62)]:
        for r in range(start, end + 1):
            ws.row_dimensions[r].outlineLevel = 1
    ws.sheet_properties.outlinePr.summaryBelow = False

    for row in range(1, 63):
        ws.row_dimensions[row].height = 22
    set_widths(ws, {"A": 23, "B": 4, "C": 24, "D": 4, "E": 14, "F": 22, "G": 12, "H": 18, "I": 16, "J": 14, "K": 14, "L": 14, "M": 14})
    for rng in ["A15:J62", "E5:M20"]:
        for row in ws[rng]:
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_reports(ws):
    ws["A1"] = "District Transfer Reports"
    ws.merge_cells("A1:J1")
    style_title(ws["A1"], 20)

    blocks = [
        ("A3", "Transfer Priority List", ["Priority", "KGID", "Name", "Current Station", "Years Served"]),
        ("G3", "Station-wise Strength", ["Station", "Sanctioned", "Working", "Vacancy"]),
        ("A20", "Sub-Division Report", ["Sub-Division", "Staff Count", "Due for Transfer"]),
        ("G20", "Rank-wise Strength", ["Rank", "Staff Count"]),
        ("A34", "Retirement Due", ["KGID", "Name", "Rank", "Retirement Date"]),
    ]
    for anchor, title, headers in blocks:
        col = ws[anchor].column
        row = ws[anchor].row
        ws[anchor] = title
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(headers) - 1)
        style_section(ws[anchor], COLORS["blue"])
        ws[anchor].font = Font(bold=True, color=COLORS["white"])
        for idx, header in enumerate(headers, start=col):
            ws.cell(row + 1, idx).value = header
        style_header(ws[row + 1][col - 1: col - 1 + len(headers)])

    for row in range(5, 20):
        src = row - 3
        ws[f"A{row}"] = f'=IF(\'Employee Master\'!$R{src}="Yes",\'Employee Master\'!$S{src},"")'
        ws[f"B{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$A{src})'
        ws[f"C{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$B{src})'
        ws[f"D{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$L{src})'
        ws[f"E{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$O{src})'
    stations = ["Robertsonpet PS", "Bethamangala PS", "Bangarpet PS", "Malur PS", "Kolar Town PS", "District Armed Reserve"]
    for i, station in enumerate(stations, start=5):
        ws[f"G{i}"] = station
        ws[f"H{i}"] = 0
        ws[f"I{i}"] = f'=COUNTIF(\'Employee Master\'!$L:$L,G{i})'
        ws[f"J{i}"] = f'=H{i}-I{i}'
    subdivisions = ["KGF", "Bangarpet", "Malur", "Kolar"]
    for i, sub in enumerate(subdivisions, start=22):
        ws[f"A{i}"] = sub
        ws[f"B{i}"] = f'=COUNTIF(\'Employee Master\'!$K:$K,A{i})'
        ws[f"C{i}"] = f'=COUNTIFS(\'Employee Master\'!$K:$K,A{i},\'Employee Master\'!$R:$R,"Yes")'
    ranks = ["PC", "HC", "ASI", "PSI", "CPI", "PI"]
    for i, rank in enumerate(ranks, start=22):
        ws[f"G{i}"] = rank
        ws[f"H{i}"] = f'=COUNTIF(\'Employee Master\'!$C:$C,G{i})'
    for row in range(36, 51):
        src = row - 34
        ws[f"A{row}"] = f'=IF(AND(\'Employee Master\'!$I{src}<>"",\'Employee Master\'!$I{src}<=EDATE(TODAY(),12)),\'Employee Master\'!$A{src},"")'
        ws[f"B{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$B{src})'
        ws[f"C{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$C{src})'
        ws[f"D{row}"] = f'=IF($A{row}="","",\'Employee Master\'!$I{src})'

    chart1 = BarChart()
    chart1.title = "Working Strength by Station"
    chart1.y_axis.title = "Staff"
    chart1.x_axis.title = "Station"
    chart1.add_data(Reference(ws, min_col=9, min_row=4, max_row=10), titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=7, min_row=5, max_row=10))
    chart1.height = 7
    chart1.width = 13
    ws.add_chart(chart1, "G12")

    chart2 = PieChart()
    chart2.title = "Rank-wise Strength"
    chart2.add_data(Reference(ws, min_col=8, min_row=21, max_row=27), titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=7, min_row=22, max_row=27))
    chart2.height = 7
    chart2.width = 10
    ws.add_chart(chart2, "G30")

    set_widths(ws, {"A": 16, "B": 14, "C": 22, "D": 24, "E": 14, "F": 4, "G": 26, "H": 14, "I": 14, "J": 14})


def build_guide(ws):
    ws["A1"] = "How to Use This Workbook"
    ws.merge_cells("A1:F1")
    style_title(ws["A1"])
    guide = [
        ("1", "Enter one row per employee in Employee Master."),
        ("2", "Enter every historical posting in Posting History. Keep current posting To Date blank."),
        ("3", "Use the Dashboard search boxes to find matching employees."),
        ("4", "Choose a KGID in Dashboard cell C12 to display full employee data."),
        ("5", "Use the grouped + / - controls on the Dashboard rows to collapse or expand sections."),
        ("6", "Review Posting History validation columns for missing dates or overlapping postings."),
        ("7", "Update sanctioned strength in Reports column H to calculate vacancy position."),
        ("8", "For Google Sheets, upload this XLSX to Drive and open as Google Sheets. Dynamic formulas may need Excel 365 or modern Google Sheets support."),
    ]
    ws.append(["Step", "Instruction"])
    style_header(ws[2][:2])
    for item in guide:
        ws.append(item)
    set_widths(ws, {"A": 10, "B": 110})
    for row in ws.iter_rows(min_row=3, max_row=10, min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")


def final_format(wb):
    thin = Side(style="thin", color="D9D9D9")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                alignment = copy(cell.alignment)
                alignment.vertical = "center"
                cell.alignment = alignment
        ws.sheet_view.zoomScale = 90


if __name__ == "__main__":
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    final_format(workbook)
    workbook.save(OUTPUT_FILE)

    loaded = load_workbook(OUTPUT_FILE, data_only=False)
    required = {"Dashboard", "Employee Master", "Posting History", "Reports", "Lists", "Guide"}
    missing = required - set(loaded.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")
    for sheet_name in ["Dashboard", "Employee Master", "Posting History", "Reports"]:
        if loaded[sheet_name].max_row < 2:
            raise RuntimeError(f"{sheet_name} appears empty")
    print(OUTPUT_FILE)
