from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS\outputs\District_Employee_Transfer_Database_IMPORTED.xlsx")
wb = load_workbook(path, data_only=False)
assert "Imported Emp Profiles" in wb.sheetnames
assert "Employee Master" in wb.sheetnames
assert wb["Imported Emp Profiles"].max_row == 976
assert wb["Employee Master"]["A2"].value == "98765"
assert wb["Employee Master"]["B2"].value == "Testrobu"
assert wb["Employee Master"]["L2"].value == "Belagavi Rural PS"
assert len(wb["Imported Emp Profiles"].tables) == 1
assert len(wb["Employee Master"].tables) == 1
print("imported workbook verified")
