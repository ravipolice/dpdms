from pathlib import Path
from openpyxl import load_workbook

base = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS")
workbook_path = base / "outputs" / "District_Employee_Transfer_Database_IMPORTED.xlsx"
dashboard_path = base / "outputs" / "Employee_Transfer_Dashboard_Browser.html"

wb = load_workbook(workbook_path, data_only=True, read_only=True)
master = wb["Employee Master"]
imported = wb["Imported Emp Profiles"]

districts = set()
master_count = 0
for row in master.iter_rows(min_row=2, values_only=True):
    kgid = row[0]
    district = row[9]
    if kgid:
        master_count += 1
        districts.add(str(district).strip())

source_districts = set()
source_count = 0
headers = [cell.value for cell in next(imported.iter_rows(min_row=1, max_row=1))]
district_idx = headers.index("district")
for row in imported.iter_rows(min_row=2, values_only=True):
    if row[0]:
        source_count += 1
        source_districts.add(str(row[district_idx]).strip())

assert master_count == 973, master_count
assert source_count == 973, source_count
assert districts == {"Chikkaballapura"}, districts
assert source_districts == {"Chikkaballapura"}, source_districts

html = dashboard_path.read_text(encoding="utf-8")
assert '"totalEmployees": 973' in html
assert "Belagavi city" not in html
assert "Chikkaballapura" in html
print("verified Chikkaballapura only")
