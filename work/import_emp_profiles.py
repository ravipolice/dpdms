from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS")
SOURCE = Path(r"C:\Users\ravip\Downloads\Emp Profiles.xlsx")
TEMPLATE = BASE / "outputs" / "District_Employee_Transfer_Database_Template_FIXED.xlsx"
OUTPUT = BASE / "outputs" / "District_Employee_Transfer_Database_IMPORTED.xlsx"

MAX_ROWS = 2500
RANK_ORDER = {"PI": 1, "CPI": 2, "PSI": 3, "ASI": 4, "HC": 5, "PC": 6}
KEEP_DISTRICTS = {"chikkaballapura", "chikkaballapur"}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def bool_text(value):
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return clean(value)


def iso_to_display(value):
    value = clean(value)
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime("%d-%m-%Y %H:%M")
    except ValueError:
        return value


def read_source():
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb["Emp Profiles"]
    raw_headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = []
    seen = {}
    for idx, header in enumerate(raw_headers, start=1):
        header = header or f"blank_{idx}"
        seen[header] = seen.get(header, 0) + 1
        headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if bool(record.get("isDeleted")):
            continue
        if not clean(record.get("kgid")) and not clean(record.get("name")):
            continue
        if clean(record.get("district")).lower() not in KEEP_DISTRICTS:
            continue
        rows.append(record)

    rows.sort(key=lambda r: (
        clean(r.get("district")).lower(),
        clean(r.get("station")).lower(),
        RANK_ORDER.get(clean(r.get("rank")).upper(), 99),
        clean(r.get("name")).lower(),
        clean(r.get("kgid")),
    ))
    return headers, rows


def clear_range(ws, min_row, max_row, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def write_row(ws, row_index, values):
    for col_index, value in enumerate(values, start=1):
        ws.cell(row_index, col_index).value = value


def set_formula_rows(ws, row):
    ws[f"I{row}"] = f'=IF(E{row}="","",EOMONTH(EDATE(E{row},60*12),0))'
    ws[f"N{row}"] = f'=IF(F{row}="","",ROUND(YEARFRAC(F{row},TODAY()),1))'
    ws[f"O{row}"] = f'=IF(M{row}="","",ROUND(YEARFRAC(M{row},TODAY()),1))'
    ws[f"P{row}"] = f'=IF(A{row}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$3000=A{row})*(\'Posting History\'!$E$2:$E$3000=K{row})*(\'Posting History\'!$B$2:$B$3000<>"")*(IF(\'Posting History\'!$C$2:$C$3000="",TODAY(),\'Posting History\'!$C$2:$C$3000)-\'Posting History\'!$B$2:$B$3000))/365.25,1))'
    ws[f"Q{row}"] = f'=IF(A{row}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$3000=A{row})*(\'Posting History\'!$F$2:$F$3000=J{row})*(\'Posting History\'!$B$2:$B$3000<>"")*(IF(\'Posting History\'!$C$2:$C$3000="",TODAY(),\'Posting History\'!$C$2:$C$3000)-\'Posting History\'!$B$2:$B$3000))/365.25,1))'
    ws[f"R{row}"] = f'=IF(O{row}="","",IF(O{row}>=3,"Yes","No"))'
    ws[f"S{row}"] = f'=IF(O{row}="","",IF(O{row}>=5,"High",IF(O{row}>=3,"Medium","Low")))'


def import_data():
    headers, rows = read_source()
    wb = load_workbook(TEMPLATE)
    master = wb["Employee Master"]
    history = wb["Posting History"]

    clear_range(master, 2, MAX_ROWS + 1, 20)
    clear_range(history, 2, MAX_ROWS + 1, 12)

    for idx, record in enumerate(rows[:MAX_ROWS], start=2):
        kgid = clean(record.get("kgid"))
        rank = clean(record.get("rank"))
        name = clean(record.get("name"))
        mobile1 = clean(record.get("mobile1"))
        mobile2 = clean(record.get("mobile2"))
        station = clean(record.get("station"))
        district = clean(record.get("district"))
        metal = clean(record.get("metalNumber"))
        blood = clean(record.get("bloodGroup"))
        email = clean(record.get("email"))
        approved = bool_text(record.get("isApproved"))
        admin = bool_text(record.get("isAdmin"))
        created = iso_to_display(record.get("createdAt"))
        updated = iso_to_display(record.get("updatedAt"))
        remarks = "; ".join(
            part for part in [
                f"Metal: {metal}" if metal else "",
                f"Blood: {blood}" if blood else "",
                f"Email: {email}" if email else "",
                f"Mobile2: {mobile2}" if mobile2 else "",
                f"Approved: {approved}" if approved else "",
                f"Admin: {admin}" if admin else "",
                f"Created: {created}" if created else "",
                f"Updated: {updated}" if updated else "",
            ] if part
        )
        write_row(master, idx, [
            kgid, name, rank, rank, None, None, mobile1, district, None, district,
            "", station, None, None, None, None, None, None, None, remarks
        ])
        set_formula_rows(master, idx)

        write_row(history, idx, [
            kgid, None, None, station, "", district, rank, "", None, None, None,
            "Current profile imported from Emp Profiles; posting dates not available."
        ])
        history[f"I{idx}"] = f'=IF(B{idx}="","",ROUND((IF(C{idx}="",TODAY(),C{idx})-B{idx})/365.25,1))'
        history[f"J{idx}"] = f'=IF(A{idx}="","",IF(COUNTIFS($A:$A,A{idx},$B:$B,"<="&IF(C{idx}="",TODAY(),C{idx}),$C:$C,">="&B{idx})>1,"Review","OK"))'
        history[f"K{idx}"] = f'=IF(A{idx}="","",IF(OR(B{idx}="",D{idx}="",F{idx}="",G{idx}=""),"Missing","OK"))'

    for ws_name, table_name, ref in [
        ("Employee Master", "tblEmployeeMaster", f"A1:T{MAX_ROWS + 1}"),
        ("Posting History", "tblPostingHistory", f"A1:L{MAX_ROWS + 1}"),
    ]:
        ws = wb[ws_name]
        ws.tables[table_name].ref = ref

    if "Imported Emp Profiles" in wb.sheetnames:
        del wb["Imported Emp Profiles"]
    imported = wb.create_sheet("Imported Emp Profiles", 1)
    source_headers = [
        "kgid", "name", "mobile1", "mobile2", "rank", "station", "district",
        "metalNumber", "bloodGroup", "email", "isAdmin", "isApproved",
        "createdAt", "updatedAt", "isDeleted"
    ]
    imported.append(source_headers)
    for record in rows:
        imported.append([
            clean(record.get("kgid")), clean(record.get("name")), clean(record.get("mobile1")),
            clean(record.get("mobile2")), clean(record.get("rank")), clean(record.get("station")),
            clean(record.get("district")), clean(record.get("metalNumber")), clean(record.get("bloodGroup")),
            clean(record.get("email")), bool_text(record.get("isAdmin")), bool_text(record.get("isApproved")),
            iso_to_display(record.get("createdAt")), iso_to_display(record.get("updatedAt")),
            bool_text(record.get("isDeleted")),
        ])

    for cell in imported[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in imported.columns:
        max_len = max(len(clean(c.value)) for c in col[:200])
        imported.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 36)
    imported.freeze_panes = "A2"
    table_ref = f"A1:O{len(rows) + 1}"
    table = Table(displayName="tblImportedEmpProfiles", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    imported.add_table(table)

    for ws in [master, history, imported]:
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False

    wb.save(OUTPUT)
    return len(rows)


if __name__ == "__main__":
    count = import_data()
    print(f"Imported {count} employees to {OUTPUT}")
