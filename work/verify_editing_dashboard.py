import importlib.util
import shutil
from pathlib import Path

from openpyxl import load_workbook

base = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS")
server_path = base / "outputs" / "employee_transfer_browser_dashboard" / "server.py"
source = base / "outputs" / "District_Employee_Transfer_Database_IMPORTED.xlsx"
temp = base / "work" / "edit_test_copy.xlsx"
shutil.copy2(source, temp)

spec = importlib.util.spec_from_file_location("dashboard_server", server_path)
module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(module)
module.WORKBOOK_PATH = temp
module.BACKUP_DIR = base / "work" / "edit_test_backups"

payload = {
    "kgid": "TESTEDIT001",
    "name": "Browser Test Employee",
    "rank": "PC",
    "designation": "PC",
    "mobile": "9999999999",
    "dob": "01-01-1990",
    "doa": "01-01-2020",
    "homeDistrict": "Chikkaballapura",
    "currentDistrict": "Chikkaballapura",
    "currentSubDivision": "Test Sub Division",
    "currentUnit": "Test PS",
    "currentSince": "01-01-2024",
    "sourceProfile": {
        "metalNumber": "T001",
        "bloodGroup": "O+",
        "email": "test@example.com",
        "mobile2": "",
        "isApproved": "Yes",
        "isAdmin": "No",
    },
}
result = module.save_employee(payload)
assert result["ok"]

wb = load_workbook(temp, data_only=False)
master = wb["Employee Master"]
found = None
for row in master.iter_rows(min_row=2, values_only=True):
    if row[0] == "TESTEDIT001":
        found = row
        break
assert found is not None
assert found[1] == "Browser Test Employee"
assert found[11] == "Test PS"

html = (base / "outputs" / "employee_transfer_browser_dashboard" / "index.html").read_text(encoding="utf-8")
assert "/api/employee" in html
assert "Save to Excel" in html
assert "Add Employee" in html
print("editing verified")
