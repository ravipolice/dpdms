from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS\outputs\District_Employee_Transfer_Database_Template.xlsx")
wb = load_workbook(path, data_only=False)

required = ["Dashboard", "Employee Master", "Posting History", "Reports", "Lists", "Guide"]
assert wb.sheetnames == required, wb.sheetnames

checks = {
    "dashboard_search_formula": wb["Dashboard"]["E6"].value,
    "dashboard_profile_formula": wb["Dashboard"]["C17"].value,
    "dashboard_history_formula": wb["Dashboard"]["A35"].value,
    "master_priority_formula": wb["Employee Master"]["S2"].value,
    "history_duration_formula": wb["Posting History"]["I2"].value,
    "reports_priority_formula": wb["Reports"]["A5"].value,
}
for name, value in checks.items():
    assert isinstance(value, str) and value.startswith("="), (name, value)

assert len(wb["Employee Master"].tables) == 1
assert len(wb["Posting History"].tables) == 1
assert len(wb["Dashboard"].data_validations.dataValidation) >= 3
assert len(wb["Employee Master"].data_validations.dataValidation) >= 4
assert len(wb["Posting History"].data_validations.dataValidation) >= 4
assert wb["Dashboard"].row_dimensions[16].outlineLevel == 1
assert len(wb["Reports"]._charts) >= 2

print("verified", path.stat().st_size)
