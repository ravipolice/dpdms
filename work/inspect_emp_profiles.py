from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\ravip\Downloads\Emp Profiles.xlsx")
print(path, path.exists(), path.stat().st_size if path.exists() else "")
wb = load_workbook(path, data_only=True, read_only=False)
print("sheets:", wb.sheetnames)
for ws in wb.worksheets:
    print("\nSHEET", ws.title, "rows", ws.max_row, "cols", ws.max_column)
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 8), values_only=True), start=1):
        print(r_idx, list(row[:min(ws.max_column, 20)]))
