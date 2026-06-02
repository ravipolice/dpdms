from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

path = Path(r"C:\Users\ravip\Downloads\Emp Profiles.xlsx")
wb = load_workbook(path, data_only=True, read_only=True)
ws = wb["Emp Profiles"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
district_index = headers.index("district")
is_deleted_index = headers.index("isDeleted")

counter = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[is_deleted_index]:
        continue
    district = row[district_index]
    counter[str(district).strip() if district is not None else ""] += 1

for district, count in counter.most_common():
    if "chikka" in district.lower() or "chik" in district.lower():
        print(district, count)
print("total variants", len(counter))
