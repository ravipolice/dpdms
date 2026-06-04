import sys
import os
import openpyxl
from pathlib import Path
from datetime import datetime

def parse_date(val):
    if not val:
        return None
    if isinstance(val, (datetime, datetime.date)):
        return val
    val_str = str(val).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return val

def import_data(source_path, master_path):
    print(f"Opening source file: {source_path}")
    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    src_sheet = src_wb.worksheets[0]
    src_rows = list(src_sheet.iter_rows(values_only=True))
    if len(src_rows) <= 1:
        print("Error: Source file has no data.")
        return

    src_headers = [str(h).strip() for h in src_rows[0]]
    src_kgid_idx = -1
    for idx, h in enumerate(src_headers):
        if h.lower() in ("kgid", "k.g.i.d", "employee kgid"):
            src_kgid_idx = idx
            break

    if src_kgid_idx == -1:
        print(f"Error: Could not find 'KGID' column in source sheet. Headers: {src_headers}")
        return

    print(f"Opening master file: {master_path}")
    master_wb = openpyxl.load_workbook(master_path)
    master_sheet = master_wb["Employee Master"]
    
    # Get master headers
    master_rows = list(master_sheet.iter_rows(values_only=True))
    master_headers = [str(h).strip() for h in master_rows[0]]
    master_kgid_idx = master_headers.index("KGID")
    
    # Map master header names to 1-based column indices
    m_cols = {h: idx + 1 for idx, h in enumerate(master_headers)}
    
    # Map source column index to master column index
    header_map = {}
    for src_idx, src_h in enumerate(src_headers):
        if src_idx == src_kgid_idx:
            continue
        
        matched_h = None
        for m_h in master_headers:
            if src_h.lower() == m_h.lower() or \
               (src_h.lower() == "name" and m_h.lower() == "employee name") or \
               (src_h.lower() == "date of birth" and m_h.lower() == "dob") or \
               (src_h.lower() == "date of appointment" and m_h.lower() == "appointment date") or \
               (src_h.lower() == "unit" and m_h.lower() == "current unit") or \
               (src_h.lower() == "subdivision" and m_h.lower() == "current sub-division") or \
               (src_h.lower() == "sub-division" and m_h.lower() == "current sub-division"):
                matched_h = m_h
                break
        
        if matched_h:
            header_map[src_idx] = m_cols[matched_h]

    # Index existing rows in master by KGID
    master_kgid_map = {}
    for r_idx in range(1, len(master_rows)):
        kgid_val = str(master_rows[r_idx][master_kgid_idx]).strip()
        if kgid_val and kgid_val != "None":
            master_kgid_map[kgid_val] = r_idx + 1 # 1-based row index

    added = 0
    updated = 0

    for r_idx in range(1, len(src_rows)):
        row_vals = src_rows[r_idx]
        if not row_vals:
            continue
        kgid = str(row_vals[src_kgid_idx]).strip()
        if not kgid or kgid == "None":
            continue

        dest_row = master_kgid_map.get(kgid)
        is_new = dest_row is None
        
        if is_new:
            dest_row = master_sheet.max_row + 1
            master_sheet.cell(row=dest_row, column=master_kgid_idx + 1, value=kgid)
            master_kgid_map[kgid] = dest_row
            added += 1
        else:
            updated += 1

        # Copy data
        for src_idx, dest_col in header_map.items():
            val = row_vals[src_idx]
            if val is not None and str(val).strip() != "":
                dest_header = master_headers[dest_col - 1]
                if dest_header in ("DOB", "Appointment Date", "Present Posting Date"):
                    val = parse_date(val)
                master_sheet.cell(row=dest_row, column=dest_col, value=val)

        # Set formulas for new rows
        if is_new:
            r = dest_row
            if "Retirement Date" in m_cols:
                master_sheet.cell(row=r, column=m_cols["Retirement Date"], value=f'=IF(E{r}="","",EOMONTH(EDATE(E{r},60*12),0))')
            if "Total Service Years" in m_cols:
                master_sheet.cell(row=r, column=m_cols["Total Service Years"], value=f'=IF(F{r}="","",ROUND(YEARFRAC(F{r},TODAY()),1)')
            if "Current Station Years" in m_cols:
                master_sheet.cell(row=r, column=m_cols["Current Station Years"], value=f'=IF(M{r}="","",ROUND(YEARFRAC(M{r},TODAY()),1)')
            if "Current Sub-Division Years" in m_cols:
                master_sheet.cell(row=r, column=m_cols["Current Sub-Division Years"], value=f'=IF(A{r}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$1000=A{r})*(\'Posting History\'!$E$2:$E$1000=K{r})*(\'Posting History\'!$B$2:$B$1000<>"")*(IF(\'Posting History\'!$C$2:$C$1000="",TODAY(),\'Posting History\'!$C$2:$C$1000)-\'Posting History\'!$B$2:$B$1000))/365.25,1))')
            if "District Service Years" in m_cols:
                master_sheet.cell(row=r, column=m_cols["District Service Years"], value=f'=IF(A{r}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$1000=A{r})*(\'Posting History\'!$F$2:$F$1000=J{r})*(\'Posting History\'!$B$2:$B$1000<>"")*(IF(\'Posting History\'!$C$2:$C$1000="",TODAY(),\'Posting History\'!$C$2:$C$1000)-\'Posting History\'!$B$2:$B$1000))/365.25,1))')
            if "Transfer Eligible" in m_cols:
                master_sheet.cell(row=r, column=m_cols["Transfer Eligible"], value=f'=IF(O{r}="","",IF(O{r}>=3,"Yes","No"))')
            if "Priority" in m_cols:
                master_sheet.cell(row=r, column=m_cols["Priority"], value=f'=IF(O{r}="","",IF(O{r}>=5,"High",IF(O{r}>=3,"Medium","Low")))')

    master_wb.save(master_path)
    print(f"\nImport Finished Successfully!\nAdded: {added} new employees\nUpdated: {updated} existing records\nSaved: {master_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_local_excel.py <path_to_downloaded_source.xlsx>")
        sys.exit(1)
    
    src = sys.argv[1]
    mst = Path("outputs/District_Employee_Transfer_Database_IMPORTED.xlsx")
    if not mst.exists():
        mst = Path("outputs/District_Employee_Transfer_Database_Template_FIXED.xlsx")
        
    if not mst.exists():
        print("Error: Master excel file not found in outputs/ directory.")
        sys.exit(1)
        
    import_data(src, str(mst))
