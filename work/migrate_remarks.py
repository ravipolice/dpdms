import openpyxl
from pathlib import Path

def parse_remarks(remarks_str):
    if not remarks_str:
        return {}, ""
    
    parts = str(remarks_str).split(";")
    profile_data = {}
    remaining_parts = []
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        if ":" in part:
            k, v = part.split(":", 1)
            k_clean = k.strip().lower()
            v_clean = v.strip()
            
            # Map known keys
            if k_clean in ("metal", "metal number", "metalno"):
                profile_data["metalNumber"] = v_clean
            elif k_clean in ("blood", "blood group", "bloodgroup"):
                profile_data["bloodGroup"] = v_clean
            elif k_clean == "email":
                profile_data["email"] = v_clean
            elif k_clean == "approved":
                profile_data["isApproved"] = v_clean
            elif k_clean == "admin":
                profile_data["isAdmin"] = v_clean
            elif k_clean == "created":
                profile_data["createdAt"] = v_clean
            elif k_clean == "updated":
                profile_data["updatedAt"] = v_clean
            else:
                remaining_parts.append(part)
        else:
            remaining_parts.append(part)
            
    remaining_remarks = "; ".join(remaining_parts).strip()
    return profile_data, remaining_remarks

def run_migration():
    db_path = Path("outputs/District_Employee_Transfer_Database_IMPORTED.xlsx")
    if not db_path.exists():
        db_path = Path("outputs/District_Employee_Transfer_Database_Template_FIXED.xlsx")
        
    if not db_path.exists():
        print("Master file not found.")
        return

    print(f"Loading {db_path}...")
    wb = openpyxl.load_workbook(db_path)
    
    master_sheet = wb["Employee Master"]
    imported_sheet = wb["Imported Emp Profiles"] if "Imported Emp Profiles" in wb.sheetnames else None
    
    if not imported_sheet:
        print("Imported Emp Profiles sheet not found. Creating it...")
        imported_sheet = wb.create_sheet("Imported Emp Profiles")
        imported_sheet.append([
            "kgid", "name", "mobile1", "mobile2", "rank", "station", "district",
            "metalNumber", "bloodGroup", "email", "isAdmin", "isApproved",
            "createdAt", "updatedAt", "isDeleted", "photoUrl", "lockedFields",
            "caste", "subCaste", "height", "weight", "familyDetails"
        ])
    
    # Read headers
    m_headers = [str(cell.value).strip() for cell in master_sheet[1]]
    kgid_col = m_headers.index("KGID") + 1
    name_col = m_headers.index("Employee Name") + 1
    remarks_col = m_headers.index("Remarks") + 1
    
    imp_headers = [str(cell.value).strip() for cell in imported_sheet[1]]
    imp_kgid_col = imp_headers.index("kgid") + 1
    imp_name_col = imp_headers.index("name") + 1
    
    # Map imported headers to column indices
    imp_cols = {h: idx + 1 for idx, h in enumerate(imp_headers)}
    
    # Index existing imported rows by KGID
    imp_kgid_row_map = {}
    for r in range(2, imported_sheet.max_row + 1):
        kgid_val = str(imported_sheet.cell(r, imp_kgid_col).value).strip()
        if kgid_val and kgid_val != "None":
            imp_kgid_row_map[kgid_val] = r

    migrated_count = 0
    cleaned_count = 0
    
    # Traverse Employee Master rows
    for r in range(2, master_sheet.max_row + 1):
        kgid = str(master_sheet.cell(r, kgid_col).value).strip()
        name = str(master_sheet.cell(r, name_col).value).strip()
        remarks_val = master_sheet.cell(r, remarks_col).value
        
        if not kgid or kgid == "None":
            continue
            
        profile_data, clean_remarks = parse_remarks(remarks_val)
        
        # If we extracted any profile details, merge them into the profile sheet
        if profile_data:
            dest_row = imp_kgid_row_map.get(kgid)
            if not dest_row:
                dest_row = imported_sheet.max_row + 1
                imported_sheet.cell(dest_row, imp_kgid_col, value=kgid)
                imported_sheet.cell(dest_row, imp_name_col, value=name)
                imp_kgid_row_map[kgid] = dest_row
                
            for k, v in profile_data.items():
                if k in imp_cols:
                    # Only overwrite if target is empty, or if we want to overwrite '??' with valid email, etc.
                    current_val = imported_sheet.cell(dest_row, imp_cols[k]).value
                    if not current_val or str(current_val).strip() in ("", "None", "??"):
                        imported_sheet.cell(dest_row, imp_cols[k], value=v)
            
            migrated_count += 1

    wb.save(db_path)
    print(f"\nMigration Completed!")
    print(f"Migrated profile tokens for {migrated_count} employees.")
    print("Original 'Remarks' column in 'Employee Master' was left untouched.")

if __name__ == "__main__":
    run_migration()
