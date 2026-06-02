from pathlib import Path
from zipfile import ZipFile
from openpyxl import load_workbook

path = Path(r"C:\Users\ravip\AndroidStudioProjects\DPDMS\outputs\District_Employee_Transfer_Database_Template_FIXED.xlsx")
wb = load_workbook(path, data_only=False)

required = ["Dashboard", "Employee Master", "Posting History", "Reports", "Lists", "Guide"]
assert wb.sheetnames == required
assert wb["Dashboard"]["E6"].value.startswith("=")
assert wb["Dashboard"]["C17"].value.startswith("=")
assert wb["Dashboard"]["A35"].value.startswith("=")
assert wb["Reports"]["A5"].value.startswith("=")
assert len(wb["Employee Master"].tables) == 1
assert len(wb["Posting History"].tables) == 1
assert len(wb["Reports"]._charts) >= 2

with ZipFile(path) as zf:
    names = zf.namelist()
    assert "[Content_Types].xml" in names
    combined = "\n".join(
        zf.read(name).decode("utf-8", errors="ignore")
        for name in names
        if name.endswith(".xml")
    )
    assert "FILTER(" not in combined
    assert "XLOOKUP(" not in combined
    assert "CHOOSECOLS(" not in combined
    assert "UNIQUE(" not in combined

print("verified fixed workbook", path.stat().st_size)
