from __future__ import annotations

import json
import mimetypes
import shutil
from datetime import date, datetime
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.parse import parse_qs

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
IMPORTED_WORKBOOK_PATH = APP_DIR.parent / "District_Employee_Transfer_Database_IMPORTED.xlsx"
FIXED_WORKBOOK_PATH = APP_DIR.parent / "District_Employee_Transfer_Database_Template_FIXED.xlsx"
WORKBOOK_PATH = IMPORTED_WORKBOOK_PATH if IMPORTED_WORKBOOK_PATH.exists() else FIXED_WORKBOOK_PATH
BACKUP_DIR = APP_DIR.parent / "backups"
MAX_BACKUPS = 30  # Keep only the last 30 rolling backups


def backup_workbook(label: str = "") -> Path | None:
    """Create a timestamped backup of the workbook and prune old backups.
    Returns the path of the new backup file, or None if workbook not found.
    """
    if not WORKBOOK_PATH.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{label}" if label else ""
    dest = BACKUP_DIR / f"{WORKBOOK_PATH.stem}{suffix}_{stamp}.xlsx"
    shutil.copy2(WORKBOOK_PATH, dest)
    # Prune: keep only the most recent MAX_BACKUPS files
    backups = sorted(BACKUP_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in backups[MAX_BACKUPS:]:
        try:
            old.unlink()
        except OSError:
            pass
    return dest


MASTER_HEADERS = [
    "KGID", "Employee Name", "Rank", "Designation", "DOB", "Appointment Date",
    "Mobile", "Home District", "Retirement Date", "Current District",
    "Current Sub-Division", "Current Unit", "Present Posting Date",
    "Total Service Years", "Current Station Years", "Current Sub-Division Years",
    "District Service Years", "Transfer Eligible", "Priority", "Remarks"
]

IMPORTED_HEADERS = [
    "kgid", "name", "mobile1", "mobile2", "rank", "station", "district",
    "metalNumber", "bloodGroup", "email", "isAdmin", "isApproved",
    "createdAt", "updatedAt", "isDeleted"
]

STATION_SUBDIVISION_MAP = {
    # Chikkaballapura Subdivision
    "Chikkaballapura Town PS": "Chikkaballapura",
    "Chikkaballapura Rural PS": "Chikkaballapura",
    "Chikkaballapura Women PS": "Chikkaballapura",
    "Chikkaballapura Traffic PS": "Chikkaballapura",
    "Gowribidanur Town PS": "Chikkaballapura",
    "Gowribidanur Rural PS": "Chikkaballapura",
    "Manchenahalli PS": "Chikkaballapura",
    "Gudibanda PS": "Chikkaballapura",
    "Nandigiridhama PS": "Chikkaballapura",
    "Peresandra PS": "Chikkaballapura",
    "Sdpo Cbpura": "Chikkaballapura",
    "CPI Office Cbpura": "Chikkaballapura",
    "CPI Off Gudibande": "Chikkaballapura",
    "CPI Off Gowribidanur": "Chikkaballapura",
    "Dysp Office Cbp": "Chikkaballapura",
    "DPO": "Chikkaballapura",
    "Control Room": "Chikkaballapura",
    "DSB": "Chikkaballapura",
    "Cen PS": "Chikkaballapura",
    "DAR Chikkaballapura": "Chikkaballapura",
    "Smmc": "Chikkaballapura",
    "DCRB": "Chikkaballapura",

    # Chintamani Subdivision
    "Chintamani Town PS": "Chintamani",
    "Chintamani Rural PS": "Chintamani",
    "Shidlagatta Town PS": "Chintamani",
    "Shidlaghatta Town PS": "Chintamani",
    "Sidlaghatta Town PS": "Chintamani",
    "Sidlagatta Town PS": "Chintamani",
    "Shidlaghatta Rural PS": "Chintamani",
    "Shidlagatta Rural PS": "Chintamani",
    "Sidlaghatta Rural PS": "Chintamani",
    "Sidlagatta Rural PS": "Chintamani",
    "Bagepalli PS": "Chintamani",
    "Chelur PS": "Chintamani",
    "Pathapalya PS": "Chintamani",
    "Dibburahalli PS": "Chintamani",
    "Batlahalli PS": "Chintamani",
    "Kencharlahalli PS": "Chintamani",
    "Dysp Off Cmi": "Chintamani",
    "CPI Office Cheluru": "Chintamani",
    "CPI Off Kencharlahalli": "Chintamani",
    "CPI Off Shidlaghatta": "Chintamani",
    "CPI Off Sidlaghatta": "Chintamani",
}


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def normalize_station(value):
    n = as_text(value).strip()
    nu = n.upper()
    if nu == "DSB":
        return "DSB"
    if nu == "DCRB":
        return "DCRB"
    if nu == "DPO":
        return "DPO"
    return n


def parse_date(value):
    value = as_text(value).strip()
    if not value:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    return value


def display_date(value):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    if value is None:
        return ""
    return str(value)


def years_between(start, end=None):
    if not isinstance(start, (date, datetime)):
        return ""
    if isinstance(start, datetime):
        start = start.date()
    end = end or date.today()
    if isinstance(end, datetime):
        end = end.date()
    return round((end - start).days / 365.25, 1)


def retirement_date(dob):
    if not isinstance(dob, (date, datetime)):
        return ""
    if isinstance(dob, datetime):
        dob = dob.date()
    year = dob.year + 60
    month = dob.month
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return next_month.replace(day=1).toordinal() - 1


def retirement_display(dob):
    ordinal = retirement_date(dob)
    if not ordinal:
        return ""
    return date.fromordinal(ordinal).strftime("%d-%m-%Y")


def usable_date(value):
    return value if isinstance(value, (date, datetime)) else None


def sheet_rows(ws):
    headers = [as_text(cell.value).strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        item = {}
        for header, value in zip(headers, row):
            item[header] = value
        rows.append(item)
    return rows


def load_data():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=True)
    master = sheet_rows(wb["Employee Master"])
    history = sheet_rows(wb["Posting History"])
    imported_rows = sheet_rows(wb["Imported Emp Profiles"]) if "Imported Emp Profiles" in wb.sheetnames else []
    source_by_kgid = {as_text(row.get("kgid")).strip(): row for row in imported_rows if as_text(row.get("kgid")).strip()}

    transfer_requests = []
    if "Transfer Requests" in wb.sheetnames:
        for tr in sheet_rows(wb["Transfer Requests"]):
            tr_kgid = as_text(tr.get("KGID")).strip()
            if tr_kgid:
                transfer_requests.append({
                    "kgid": tr_kgid,
                    "name": as_text(tr.get("Employee Name")),
                    "rank": as_text(tr.get("Rank")),
                    "currentStation": as_text(tr.get("Current Station")),
                    "preference1": as_text(tr.get("Preference 1")),
                    "preference2": as_text(tr.get("Preference 2")),
                    "preference3": as_text(tr.get("Preference 3")),
                    "preference4": as_text(tr.get("Preference 4")),
                    "preference5": as_text(tr.get("Preference 5")),
                    "transferCategory": as_text(tr.get("Transfer Category")),
                    "reason": as_text(tr.get("Reason")),
                    "remarks": as_text(tr.get("Remarks")),
                    "applicationDate": as_text(tr.get("Application Date")),
                    "status": as_text(tr.get("Status") or "Pending"),
                    "approvedStation": as_text(tr.get("Approved Station") or "")
                })

    history_by_kgid = {}
    for item in history:
        kgid = as_text(item.get("KGID")).strip()
        if not kgid:
            continue
        from_date = item.get("From Date")
        to_date = item.get("To Date")
        item["_fromDisplay"] = display_date(from_date)
        item["_toDisplay"] = display_date(to_date) or "Present"
        item["_durationYears"] = years_between(from_date, to_date)
        history_by_kgid.setdefault(kgid, []).append(item)

    employees = []
    for employee in master:
        kgid = as_text(employee.get("KGID")).strip()
        if not kgid:
            continue
        dob = employee.get("DOB")
        doa = employee.get("Appointment Date")
        current_since = employee.get("Present Posting Date")
        total_service = years_between(doa)
        current_station_years = years_between(current_since)
        priority = "High" if current_station_years != "" and current_station_years >= 5 else "Medium" if current_station_years != "" and current_station_years >= 3 else "Low"
        eligible = "Yes" if current_station_years != "" and current_station_years >= 3 else "No"
        postings = history_by_kgid.get(kgid, [])
        source = source_by_kgid.get(kgid, {})
        longest = max([p.get("_durationYears") or 0 for p in postings], default=0)
        retirement_value = usable_date(employee.get("Retirement Date"))
        category = as_text(employee.get("Category")).strip()
        type_of_transfer = as_text(employee.get("Type of Transfer")).strip() or "Regular"
        current_subdiv_raw = as_text(employee.get("Current Sub-Division")).strip()
        unit = normalize_station(employee.get("Current Unit"))
        if not current_subdiv_raw:
            current_subdiv_raw = STATION_SUBDIVISION_MAP.get(unit, "")
        current_district_raw = as_text(employee.get("Current District")).strip()
        subdiv_years = round(sum(
            (p.get("_durationYears") or 0) for p in postings
            if current_subdiv_raw and (as_text(p.get("Sub-Division")).strip() or STATION_SUBDIVISION_MAP.get(normalize_station(p.get("Police Station / Unit")), "")) == current_subdiv_raw
        ), 1) if current_subdiv_raw else ""
        district_years_val = round(sum(
            (p.get("_durationYears") or 0) for p in postings
            if current_district_raw and as_text(p.get("District")).strip() == current_district_raw
        ), 1) if current_district_raw else ""
        ret_date_str = display_date(retirement_value) or retirement_display(dob)
        retirement_months_left = ""
        if ret_date_str:
            try:
                rd = datetime.strptime(ret_date_str, "%d-%m-%Y").date()
                today_d = date.today()
                retirement_months_left = (rd.year - today_d.year) * 12 + (rd.month - today_d.month)
            except Exception:
                pass
        employee_out = {
            "kgid": kgid,
            "name": as_text(employee.get("Employee Name")),
            "rank": as_text(employee.get("Rank")),
            "designation": as_text(employee.get("Designation")),
            "dob": display_date(dob),
            "doa": display_date(doa),
            "mobile": as_text(employee.get("Mobile")),
            "homeDistrict": as_text(employee.get("Home District")),
            "retirementDate": display_date(retirement_value) or retirement_display(dob),
            "currentDistrict": as_text(employee.get("Current District")),
            "currentSubDivision": current_subdiv_raw,
            "currentUnit": unit,
            "currentSince": display_date(current_since),
            "totalServiceYears": total_service,
            "currentStationYears": current_station_years,
            "transferEligible": eligible,
            "priority": priority,
            "longestPostingYears": longest,
            "postings": [
                {
                    "from": item["_fromDisplay"],
                    "to": item["_toDisplay"],
                    "station": normalize_station(item.get("Police Station / Unit")),
                    "subDivision": as_text(item.get("Sub-Division")).strip() or STATION_SUBDIVISION_MAP.get(normalize_station(item.get("Police Station / Unit")), ""),
                    "district": as_text(item.get("District")),
                    "rank": as_text(item.get("Rank Held")),
                    "orderNumber": as_text(item.get("Order Number")),
                    "durationYears": item["_durationYears"],
                }
                for item in postings
            ],
            "sourceProfile": {
                "metalNumber": as_text(source.get("metalNumber")),
                "bloodGroup": as_text(source.get("bloodGroup")),
                "email": as_text(source.get("email")),
                "mobile2": as_text(source.get("mobile2")),
                "isAdmin": as_text(source.get("isAdmin")),
                "isApproved": as_text(source.get("isApproved")),
                "createdAt": as_text(source.get("createdAt")),
                "updatedAt": as_text(source.get("updatedAt")),
                "height": as_text(source.get("height")),
                "weight": as_text(source.get("weight")),
                "caste": as_text(source.get("caste")),
                "subCaste": as_text(source.get("subCaste")),
                "familyDetails": as_text(source.get("familyDetails")),
                "photoUrl": as_text(source.get("photoUrl") or source.get("photo") or ""),
                "lockedFields": as_text(source.get("lockedFields") or ""),
            },
            "category": category,
            "typeOfTransfer": type_of_transfer,
            "subDivisionYears": subdiv_years,
            "districtServiceYears": district_years_val,
            "retirementMonthsLeft": retirement_months_left,
        }
        employees.append(employee_out)

    stations = sorted({e["currentUnit"] for e in employees if e["currentUnit"]})
    ranks = sorted({e["rank"] for e in employees if e["rank"]})
    subdivisions = sorted({e["currentSubDivision"] for e in employees if e["currentSubDivision"]})

    summary = {
        "totalEmployees": len(employees),
        "transferDue": sum(1 for e in employees if e["transferEligible"] == "Yes"),
        "highPriority": sum(1 for e in employees if e["priority"] == "High"),
        "retirementDue": sum(1 for e in employees if e["retirementDate"]),
        "retirementDue12m": sum(
            1 for e in employees
            if isinstance(e["retirementMonthsLeft"], int) and 0 <= e["retirementMonthsLeft"] <= 12
        ),
        "stations": [{"name": station, "working": sum(1 for e in employees if e["currentUnit"] == station)} for station in stations],
        "ranks": [{"name": rank, "working": sum(1 for e in employees if e["rank"] == rank)} for rank in ranks],
        "subdivisions": [{"name": sub, "working": sum(1 for e in employees if e["currentSubDivision"] == sub)} for sub in subdivisions],
    }

    return {
        "workbook": str(WORKBOOK_PATH),
        "editable": True,
        "lastLoaded": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
        "employees": employees,
        "summary": summary,
        "transferRequests": transfer_requests,
    }


def headers_for(ws):
    return [as_text(cell.value).strip() for cell in ws[1]]


def find_row_by_value(ws, header_name, value):
    headers = headers_for(ws)
    if header_name not in headers:
        return None
    col = headers.index(header_name) + 1
    target = as_text(value).strip()
    for row in range(2, ws.max_row + 1):
        if as_text(ws.cell(row, col).value).strip() == target:
            return row
    return None


def find_rows_by_value(ws, header_name, value):
    headers = headers_for(ws)
    if header_name not in headers:
        return []
    col = headers.index(header_name) + 1
    target = as_text(value).strip()
    if not target:
        return []
    rows = []
    for row in range(2, ws.max_row + 1):
        if as_text(ws.cell(row, col).value).strip() == target:
            rows.append(row)
    return rows


def first_blank_row(ws, key_col=1):
    for row in range(2, ws.max_row + 2):
        if not as_text(ws.cell(row, key_col).value).strip():
            return row
    return ws.max_row + 1


def set_master_formulas(ws, row):
    ws[f"I{row}"] = f'=IF(E{row}="","",EOMONTH(EDATE(E{row},60*12),0))'
    ws[f"N{row}"] = f'=IF(F{row}="","",ROUND(YEARFRAC(F{row},TODAY()),1))'
    ws[f"O{row}"] = f'=IF(M{row}="","",ROUND(YEARFRAC(M{row},TODAY()),1))'
    ws[f"P{row}"] = f'=IF(A{row}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$3000=A{row})*(\'Posting History\'!$E$2:$E$3000=K{row})*(\'Posting History\'!$B$2:$B$3000<>"")*(IF(\'Posting History\'!$C$2:$C$3000="",TODAY(),\'Posting History\'!$C$2:$C$3000)-\'Posting History\'!$B$2:$B$3000))/365.25,1))'
    ws[f"Q{row}"] = f'=IF(A{row}="","",ROUND(SUMPRODUCT((\'Posting History\'!$A$2:$A$3000=A{row})*(\'Posting History\'!$F$2:$F$3000=J{row})*(\'Posting History\'!$B$2:$B$3000<>"")*(IF(\'Posting History\'!$C$2:$C$3000="",TODAY(),\'Posting History\'!$C$2:$C$3000)-\'Posting History\'!$B$2:$B$3000))/365.25,1))'
    ws[f"R{row}"] = f'=IF(O{row}="","",IF(O{row}>=3,"Yes","No"))'
    ws[f"S{row}"] = f'=IF(O{row}="","",IF(O{row}>=5,"High",IF(O{row}>=3,"Medium","Low")))'


def update_table_refs(wb):
    table_specs = [
        ("Employee Master", "tblEmployeeMaster"),
        ("Posting History", "tblPostingHistory"),
        ("Imported Emp Profiles", "tblImportedEmpProfiles"),
    ]
    for sheet_name, table_name in table_specs:
        if sheet_name in wb.sheetnames and table_name in wb[sheet_name].tables:
            ws = wb[sheet_name]
            last_row = max(2, ws.max_row)
            last_col = get_column_letter(ws.max_column)
            ws.tables[table_name].ref = f"A1:{last_col}{last_row}"


def save_employee(payload):
    kgid = as_text(payload.get("kgid")).strip()
    if not kgid:
        raise ValueError("KGID is required.")
    name = as_text(payload.get("name")).strip()
    if not name:
        raise ValueError("Employee name is required.")

    backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    master = wb["Employee Master"]
    history = wb["Posting History"]
    imported = wb["Imported Emp Profiles"] if "Imported Emp Profiles" in wb.sheetnames else None

    original_kgid = as_text(payload.get("originalKgid")).strip()
    is_add = not original_kgid
    mobile = as_text(payload.get("mobile")).strip()
    source_profile = payload.get("sourceProfile") or {}
    mobile2 = as_text(source_profile.get("mobile2")).strip()
    email = as_text(source_profile.get("email")).strip().lower()

    kgid_rows = find_rows_by_value(master, "KGID", kgid)
    if is_add and kgid_rows:
        existing_name = as_text(master.cell(kgid_rows[0], 2).value).strip()
        raise ValueError(f"Duplicate KGID found: {kgid} already belongs to {existing_name or 'another employee'}.")
    if not is_add and original_kgid != kgid and kgid_rows:
        existing_name = as_text(master.cell(kgid_rows[0], 2).value).strip()
        raise ValueError(f"Cannot change KGID to {kgid}; it already belongs to {existing_name or 'another employee'}.")

    for mobile_value in [mobile, mobile2]:
        if not mobile_value:
            continue
        for header in ["Mobile"]:
            for row in find_rows_by_value(master, header, mobile_value):
                row_kgid = as_text(master.cell(row, 1).value).strip()
                if (is_add and row_kgid != kgid) or (not is_add and row_kgid not in {original_kgid, kgid}):
                    existing_name = as_text(master.cell(row, 2).value).strip()
                    raise ValueError(f"Duplicate mobile number found: {mobile_value} already belongs to {existing_name or row_kgid}.")
        if imported is not None:
            for header in ["mobile1", "mobile2"]:
                for row in find_rows_by_value(imported, header, mobile_value):
                    imported_headers = headers_for(imported)
                    row_kgid = as_text(imported.cell(row, imported_headers.index("kgid") + 1).value).strip() if "kgid" in imported_headers else ""
                    if (is_add and row_kgid != kgid) or (not is_add and row_kgid not in {original_kgid, kgid}):
                        name_col = imported_headers.index("name") + 1 if "name" in imported_headers else 2
                        existing_name = as_text(imported.cell(row, name_col).value).strip()
                        raise ValueError(f"Duplicate mobile number found: {mobile_value} already belongs to {existing_name or row_kgid}.")

    if email and imported is not None:
        imported_headers = headers_for(imported)
        if "email" in imported_headers:
            email_col = imported_headers.index("email") + 1
            kgid_col = imported_headers.index("kgid") + 1 if "kgid" in imported_headers else 1
            name_col = imported_headers.index("name") + 1 if "name" in imported_headers else 2
            for row in range(2, imported.max_row + 1):
                row_email = as_text(imported.cell(row, email_col).value).strip().lower()
                if row_email != email:
                    continue
                row_kgid = as_text(imported.cell(row, kgid_col).value).strip()
                if (is_add and row_kgid != kgid) or (not is_add and row_kgid not in {original_kgid, kgid}):
                    existing_name = as_text(imported.cell(row, name_col).value).strip()
                    raise ValueError(f"Duplicate email found: {email} already belongs to {existing_name or row_kgid}.")

    master_row = find_row_by_value(master, "KGID", kgid) or first_blank_row(master)
    master_values = {
        "KGID": kgid,
        "Employee Name": name,
        "Rank": as_text(payload.get("rank")).strip(),
        "Designation": as_text(payload.get("designation")).strip() or as_text(payload.get("rank")).strip(),
        "DOB": parse_date(payload.get("dob")),
        "Appointment Date": parse_date(payload.get("doa")),
        "Mobile": as_text(payload.get("mobile")).strip(),
        "Home District": as_text(payload.get("homeDistrict")).strip(),
        "Current District": as_text(payload.get("currentDistrict")).strip() or "Chikkaballapura",
        "Current Sub-Division": as_text(payload.get("currentSubDivision")).strip(),
        "Current Unit": normalize_station(payload.get("currentUnit")),
        "Present Posting Date": parse_date(payload.get("currentSince")),
        "Remarks": as_text(payload.get("remarks")).strip(),
        "Category": as_text(payload.get("category")).strip(),
        "Type of Transfer": as_text(payload.get("typeOfTransfer")).strip() or "Regular",
    }
    master_headers = headers_for(master)
    if "Type of Transfer" not in master_headers:
        new_col = len(master_headers) + 1
        master.cell(1, new_col).value = "Type of Transfer"
        master_headers.append("Type of Transfer")
    for header, value in master_values.items():
        if header in master_headers:
            master.cell(master_row, master_headers.index(header) + 1).value = value
    set_master_formulas(master, master_row)

    if "postings" in payload:
        # Delete all existing postings in Posting History for original_kgid / kgid
        search_kgid = original_kgid if original_kgid else kgid
        history_rows = sorted(find_rows_by_value(history, "KGID", search_kgid), reverse=True)
        for r in history_rows:
            history.delete_rows(r)

        history_headers = headers_for(history)
        # Save all postings from payload
        for p in payload.get("postings", []):
            history_row = first_blank_row(history)
            p_from = parse_date(p.get("from"))
            p_to = parse_date(p.get("to")) if p.get("to") and p.get("to") != "Present" else None
            
            history_values = {
                "KGID": kgid,
                "From Date": p_from,
                "To Date": p_to,
                "Police Station / Unit": normalize_station(p.get("station")),
                "Sub-Division": as_text(p.get("subDivision")).strip(),
                "District": as_text(p.get("district")).strip() or "Chikkaballapura",
                "Rank Held": as_text(p.get("rank")).strip(),
                "Order Number": as_text(p.get("orderNumber")).strip(),
                "Notes": "Maintained from browser dashboard.",
            }
            for header, value in history_values.items():
                if header in history_headers:
                    history.cell(history_row, history_headers.index(header) + 1).value = value
            history[f"I{history_row}"] = f'=IF(B{history_row}="","",ROUND((IF(C{history_row}="",TODAY(),C{history_row})-B{history_row})/365.25,1))'
            history[f"J{history_row}"] = f'=IF(A{history_row}="","",IF(COUNTIFS($A:$A,A{history_row},$B:$B,"<="&IF(C{history_row}="",TODAY(),C{history_row}),$C:$C,">="&B{history_row})>1,"Review","OK"))'
            history[f"K{history_row}"] = f'=IF(A{history_row}="","",IF(OR(B{history_row}="",D{history_row}="",F{history_row}="",G{history_row}=""),"Missing","OK"))'
    else:
        # Fallback to old single-posting logic
        history_row = find_row_by_value(history, "KGID", kgid) or first_blank_row(history)
        history_values = {
            "KGID": kgid,
            "From Date": parse_date(payload.get("currentSince")),
            "To Date": None,
            "Police Station / Unit": normalize_station(payload.get("currentUnit")),
            "Sub-Division": as_text(payload.get("currentSubDivision")).strip(),
            "District": as_text(payload.get("currentDistrict")).strip() or "Chikkaballapura",
            "Rank Held": as_text(payload.get("rank")).strip(),
            "Order Number": as_text(payload.get("orderNumber")).strip(),
            "Notes": "Current profile maintained from browser dashboard.",
        }
        history_headers = headers_for(history)
        for header, value in history_values.items():
            if header in history_headers:
                history.cell(history_row, history_headers.index(header) + 1).value = value
        history[f"I{history_row}"] = f'=IF(B{history_row}="","",ROUND((IF(C{history_row}="",TODAY(),C{history_row})-B{history_row})/365.25,1))'
        history[f"J{history_row}"] = f'=IF(A{history_row}="","",IF(COUNTIFS($A:$A,A{history_row},$B:$B,"<="&IF(C{history_row}="",TODAY(),C{history_row}),$C:$C,">="&B{history_row})>1,"Review","OK"))'
        history[f"K{history_row}"] = f'=IF(A{history_row}="","",IF(OR(B{history_row}="",D{history_row}="",F{history_row}="",G{history_row}=""),"Missing","OK"))'

    source = payload.get("sourceProfile") or {}
    if imported is not None:
        imported_row = find_row_by_value(imported, "kgid", kgid) or first_blank_row(imported)
        imported_values = {
            "kgid": kgid,
            "name": name,
            "mobile1": as_text(payload.get("mobile")).strip(),
            "mobile2": as_text(source.get("mobile2")).strip(),
            "rank": as_text(payload.get("rank")).strip(),
            "station": normalize_station(payload.get("currentUnit")),
            "district": as_text(payload.get("currentDistrict")).strip() or "Chikkaballapura",
            "metalNumber": as_text(source.get("metalNumber")).strip(),
            "bloodGroup": as_text(source.get("bloodGroup")).strip(),
            "email": as_text(source.get("email")).strip(),
            "isAdmin": as_text(source.get("isAdmin")).strip(),
            "isApproved": as_text(source.get("isApproved")).strip() or "Yes",
            "createdAt": as_text(source.get("createdAt")).strip(),
            "updatedAt": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "isDeleted": "No",
            "height": as_text(source.get("height")).strip(),
            "weight": as_text(source.get("weight")).strip(),
            "caste": as_text(source.get("caste")).strip(),
            "subCaste": as_text(source.get("subCaste")).strip(),
            "familyDetails": as_text(source.get("familyDetails")).strip(),
            "photoUrl": as_text(source.get("photoUrl")).strip(),
            "lockedFields": as_text(source.get("lockedFields") or "").strip(),
        }
        imported_headers = headers_for(imported)
        for col_name in ["email", "photoUrl", "lockedFields", "caste", "subCaste", "height", "weight", "familyDetails"]:
            if col_name not in imported_headers:
                new_col = len(imported_headers) + 1
                imported.cell(1, new_col).value = col_name
                imported_headers.append(col_name)
        for header, value in imported_values.items():
            if header in imported_headers:
                imported.cell(imported_row, imported_headers.index(header) + 1).value = value

    update_table_refs(wb)
    wb.save(WORKBOOK_PATH)
    return {"ok": True, "message": f"Saved {name} ({kgid})", "backupFolder": str(BACKUP_DIR)}


def delete_employee(kgid):
    if not kgid:
        raise ValueError("KGID is required.")

    backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    master = wb["Employee Master"]
    history = wb["Posting History"]
    imported = wb["Imported Emp Profiles"] if "Imported Emp Profiles" in wb.sheetnames else None

    master_row = find_row_by_value(master, "KGID", kgid)
    if not master_row:
        raise ValueError(f"Employee {kgid} not found in Employee Master.")

    master_headers = headers_for(master)
    name_col = master_headers.index("Employee Name") + 1 if "Employee Name" in master_headers else 2
    name = as_text(master.cell(master_row, name_col).value).strip()

    master.delete_rows(master_row)

    history_rows = sorted(find_rows_by_value(history, "KGID", kgid), reverse=True)
    for row in history_rows:
        history.delete_rows(row)

    if imported is not None:
        imp_row = find_row_by_value(imported, "kgid", kgid)
        if imp_row:
            imported.delete_rows(imp_row)

    update_table_refs(wb)
    wb.save(WORKBOOK_PATH)
    return {"ok": True, "message": f"Deleted {name} ({kgid})", "backupFolder": str(BACKUP_DIR)}


def verify_payload_genuinity(payload):
    kgid = as_text(payload.get("kgid")).strip()
    dob_str = as_text(payload.get("dob")).strip()
    doa_str = as_text(payload.get("doa")).strip()
    current_since_str = as_text(payload.get("currentSince")).strip()
    postings = payload.get("postings") or []

    dob = parse_date(dob_str)
    doa = parse_date(doa_str)
    current_since = parse_date(current_since_str)
    today = date.today()

    errors = []

    # 1. DOB checks
    if dob_str:
        if not isinstance(dob, (date, datetime)):
            errors.append("Invalid Date of Birth format. Please use DD-MM-YYYY.")
        elif dob > today:
            errors.append("Date of Birth cannot be in the future.")
        else:
            age = (today - dob).days / 365.25
            if age < 18:
                errors.append("Employee must be at least 18 years old.")

    # 2. Appointment checks
    if doa_str:
        if not isinstance(doa, (date, datetime)):
            errors.append("Invalid Date of Appointment format. Please use DD-MM-YYYY.")
        elif doa > today:
            errors.append("Date of Appointment cannot be in the future.")

    if dob and doa:
        if isinstance(dob, (date, datetime)) and isinstance(doa, (date, datetime)):
            if doa < dob:
                errors.append("Date of Appointment cannot be before Date of Birth.")
            elif (doa - dob).days / 365.25 < 18:
                errors.append("Appointment Date must be at least 18 years after Date of Birth.")

    # 3. Current joining date checks
    if current_since_str:
        if not isinstance(current_since, (date, datetime)):
            errors.append("Invalid Joining Date format. Please use DD-MM-YYYY.")
        elif current_since > today:
            errors.append("Joining Date at current station cannot be in the future.")

    if doa and current_since:
        if isinstance(doa, (date, datetime)) and isinstance(current_since, (date, datetime)):
            if current_since < doa:
                errors.append("Joining Date at current station cannot be before Date of Appointment.")

    # 4. Postings checks
    parsed_postings = []
    for idx, p in enumerate(postings):
        p_from_str = as_text(p.get("from")).strip()
        p_to_str = as_text(p.get("to")).strip()
        p_station = as_text(p.get("station")).strip()
        p_from = parse_date(p_from_str)
        p_to = parse_date(p_to_str) if p_to_str and p_to_str != "Present" else None

        if p_from_str and not isinstance(p_from, (date, datetime)):
            errors.append(f"Posting #{idx+1} ({p_station}): Invalid From date format.")
        if p_to_str and p_to_str != "Present" and not isinstance(p_to, (date, datetime)):
            errors.append(f"Posting #{idx+1} ({p_station}): Invalid To date format.")

        if p_from and p_to:
            if isinstance(p_from, (date, datetime)) and isinstance(p_to, (date, datetime)):
                if p_to < p_from:
                    errors.append(f"Posting #{idx+1} ({p_station}): To date cannot be before From date.")
                if doa and isinstance(doa, (date, datetime)) and p_from < doa:
                    errors.append(f"Posting #{idx+1} ({p_station}): From date cannot be before Appointment Date.")
                if current_since and isinstance(current_since, (date, datetime)) and p_to > current_since:
                    errors.append(f"Posting #{idx+1} ({p_station}): To date cannot be after Current Unit Joining Date.")
                parsed_postings.append((p_from, p_to, p_station, idx+1))

    # Overlaps checks
    parsed_postings.sort(key=lambda x: x[0])
    for i in range(len(parsed_postings) - 1):
        curr_from, curr_to, curr_station, curr_idx = parsed_postings[i]
        next_from, next_to, next_station, next_idx = parsed_postings[i+1]
        if curr_to and next_from and curr_to > next_from:
            errors.append(f"Overlap detected between postings: '{curr_station}' (ends {curr_to.strftime('%d-%m-%Y')}) and '{next_station}' (starts {next_from.strftime('%d-%m-%Y')}).")

    return errors


def save_transfer_request(payload):
    kgid = as_text(payload.get("kgid")).strip()
    if not kgid:
        raise ValueError("KGID is required.")
    
    # 1. Server-side genuinity verification
    errors = verify_payload_genuinity(payload)
    if errors:
        raise ValueError("Verification failed:\n" + "\n".join(f"- {e}" for e in errors))

    name = as_text(payload.get("name")).strip()
    rank = as_text(payload.get("rank")).strip()
    current_station = as_text(payload.get("currentStation")).strip()
    pref1 = as_text(payload.get("preference1")).strip()
    pref2 = as_text(payload.get("preference2")).strip()
    pref3 = as_text(payload.get("preference3")).strip()
    pref4 = as_text(payload.get("preference4")).strip()
    pref5 = as_text(payload.get("preference5")).strip()
    transfer_category = as_text(payload.get("transferCategory")).strip()
    reason = as_text(payload.get("reason")).strip()
    remarks = as_text(payload.get("remarks")).strip()
    status = as_text(payload.get("status") or "Pending").strip()
    apply_date = as_text(payload.get("applicationDate") or datetime.now().strftime("%d-%m-%Y")).strip()
    approved_station = as_text(payload.get("approvedStation") or "").strip()

    backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    
    # 2. Crosscheck and sync with Employee Master
    warnings = []
    if "Employee Master" in wb.sheetnames:
        master = wb["Employee Master"]
        master_headers = headers_for(master)
        master_row = find_row_by_value(master, "KGID", kgid)
        if master_row:
            def check_field(header_name, payload_val, friendly_name):
                if header_name in master_headers:
                    col_idx = master_headers.index(header_name) + 1
                    sheet_val = master.cell(master_row, col_idx).value
                    sheet_norm = as_text(sheet_val).strip()
                    pay_norm = as_text(payload_val).strip()
                    
                    if isinstance(sheet_val, (date, datetime)):
                        sheet_norm = sheet_val.strftime("%d-%m-%Y")
                    
                    parsed_pay = parse_date(pay_norm)
                    if isinstance(parsed_pay, (date, datetime)):
                        pay_norm = parsed_pay.strftime("%d-%m-%Y")
                        
                    if sheet_norm and pay_norm and sheet_norm != pay_norm:
                        warnings.append(f"{friendly_name}: entered '{pay_norm}' does not match profile '{sheet_norm}'")

            check_field("DOB", payload.get("dob"), "DOB")
            check_field("Appointment Date", payload.get("doa"), "DOA")
            check_field("Current Unit", payload.get("currentStation"), "Current Station")
            check_field("Present Posting Date", payload.get("currentSince"), "Joining Date")
            check_field("Type of Transfer", payload.get("typeOfTransfer"), "Type of Transfer")

            # Update Employee Master
            if "dob" in payload:
                master.cell(master_row, master_headers.index("DOB") + 1).value = parse_date(payload.get("dob"))
            if "doa" in payload:
                master.cell(master_row, master_headers.index("Appointment Date") + 1).value = parse_date(payload.get("doa"))
            if "currentStation" in payload:
                master.cell(master_row, master_headers.index("Current Unit") + 1).value = normalize_station(payload.get("currentStation"))
            if "currentSince" in payload:
                master.cell(master_row, master_headers.index("Present Posting Date") + 1).value = parse_date(payload.get("currentSince"))
            if "typeOfTransfer" in payload:
                if "Type of Transfer" not in master_headers:
                    new_col = len(master_headers) + 1
                    master.cell(1, new_col).value = "Type of Transfer"
                    master_headers.append("Type of Transfer")
                master.cell(master_row, master_headers.index("Type of Transfer") + 1).value = as_text(payload.get("typeOfTransfer")).strip() or "Regular"
            
            set_master_formulas(master, master_row)

    # 3. Update Posting History if provided
    if "postings" in payload and "Posting History" in wb.sheetnames:
        history = wb["Posting History"]
        history_rows = sorted(find_rows_by_value(history, "KGID", kgid), reverse=True)
        for r in history_rows:
            history.delete_rows(r)
            
        history_headers = headers_for(history)
        for p in payload.get("postings", []):
            history_row = first_blank_row(history)
            p_from = parse_date(p.get("from"))
            p_to = parse_date(p.get("to")) if p.get("to") and p.get("to") != "Present" else None
            
            history_values = {
                "KGID": kgid,
                "From Date": p_from,
                "To Date": p_to,
                "Police Station / Unit": normalize_station(p.get("station")),
                "Sub-Division": as_text(p.get("subDivision")).strip() or STATION_SUBDIVISION_MAP.get(normalize_station(p.get("station")), ""),
                "District": as_text(p.get("district")).strip() or "Chikkaballapura",
                "Rank Held": as_text(p.get("rank")).strip() or rank,
                "Order Number": as_text(p.get("orderNumber")).strip(),
                "Notes": "Maintained from transfer application.",
            }
            for header, value in history_values.items():
                if header in history_headers:
                    history.cell(history_row, history_headers.index(header) + 1).value = value
            history[f"I{history_row}"] = f'=IF(B{history_row}="","",ROUND((IF(C{history_row}="",TODAY(),C{history_row})-B{history_row})/365.25,1))'
            history[f"J{history_row}"] = f'=IF(A{history_row}="","",IF(COUNTIFS($A:$A,A{history_row},$B:$B,"<="&IF(C{history_row}="",TODAY(),C{history_row}),$C:$C,">="&B{history_row})>1,"Review","OK"))'
            history[f"K{history_row}"] = f'=IF(A{history_row}="","",IF(OR(B{history_row}="",D{history_row}="",F{history_row}="",G{history_row}=""),"Missing","OK"))'

    # 4. Save Transfer Request to Transfer Requests sheet
    if "Transfer Requests" not in wb.sheetnames:
        ws = wb.create_sheet("Transfer Requests")
        ws.append(["KGID", "Employee Name", "Rank", "Current Station", "Transfer Category", "Preference 1", "Preference 2", "Preference 3", "Preference 4", "Preference 5", "Reason", "Remarks", "Application Date", "Status", "Approved Station"])
    else:
        ws = wb["Transfer Requests"]
    
    headers = headers_for(ws)
    # Migrate: add new columns if missing
    for new_col in ["Preference 4", "Preference 5", "Transfer Category", "Approved Station"]:
        if new_col not in headers:
            ws.cell(1, len(headers) + 1).value = new_col
            headers = headers_for(ws)
    row_idx = find_row_by_value(ws, "KGID", kgid)
    if not row_idx:
        row_idx = first_blank_row(ws)
    
    values = {
        "KGID": kgid,
        "Employee Name": name,
        "Rank": rank,
        "Current Station": current_station,
        "Transfer Category": transfer_category,
        "Preference 1": pref1,
        "Preference 2": pref2,
        "Preference 3": pref3,
        "Preference 4": pref4,
        "Preference 5": pref5,
        "Reason": reason,
        "Remarks": remarks,
        "Application Date": apply_date,
        "Status": status,
        "Approved Station": approved_station
    }
    
    for header, value in values.items():
        if header in headers:
            ws.cell(row_idx, headers.index(header) + 1).value = value
            
    if "tblTransferRequests" in ws.tables:
        last_row = max(2, ws.max_row)
        last_col = get_column_letter(ws.max_column)
        ws.tables["tblTransferRequests"].ref = f"A1:{last_col}{last_row}"
    else:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        last_row = max(2, ws.max_row)
        last_col = get_column_letter(ws.max_column)
        tab = Table(displayName="tblTransferRequests", ref=f"A1:{last_col}{last_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
    update_table_refs(wb)
    wb.save(WORKBOOK_PATH)
    return {"ok": True, "message": f"Saved transfer request for {name} ({kgid})", "backupFolder": str(BACKUP_DIR), "warnings": warnings}


def delete_transfer_request(kgid):
    if not kgid:
        raise ValueError("KGID is required.")
        
    backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    if "Transfer Requests" in wb.sheetnames:
        ws = wb["Transfer Requests"]
        row_idx = find_row_by_value(ws, "KGID", kgid)
        if row_idx:
            ws.delete_rows(row_idx)
            if "tblTransferRequests" in ws.tables:
                last_row = max(2, ws.max_row)
                last_col = get_column_letter(ws.max_column)
                ws.tables["tblTransferRequests"].ref = f"A1:{last_col}{last_row}"
            wb.save(WORKBOOK_PATH)
            return {"ok": True, "message": f"Withdrew transfer request for KGID {kgid}", "backupFolder": str(BACKUP_DIR)}
            
    return {"ok": False, "error": f"Transfer request for KGID {kgid} not found."}


def duplicate_status(query):
    kgid = as_text(query.get("kgid", [""])[0]).strip()
    original_kgid = as_text(query.get("originalKgid", [""])[0]).strip()
    mobile = as_text(query.get("mobile", [""])[0]).strip()
    mobile2 = as_text(query.get("mobile2", [""])[0]).strip()
    email = as_text(query.get("email", [""])[0]).strip().lower()
    is_add = not original_kgid

    wb = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    master = wb["Employee Master"]
    imported = wb["Imported Emp Profiles"] if "Imported Emp Profiles" in wb.sheetnames else None
    issues = []

    def same_record(row_kgid):
        return (not is_add) and row_kgid in {original_kgid, kgid}

    master_headers = headers_for(master)
    master_cols = {header: idx + 1 for idx, header in enumerate(master_headers)}
    for row in range(2, master.max_row + 1):
        row_kgid = as_text(master.cell(row, master_cols.get("KGID", 1)).value).strip()
        row_name = as_text(master.cell(row, master_cols.get("Employee Name", 2)).value).strip()
        row_mobile = as_text(master.cell(row, master_cols.get("Mobile", 7)).value).strip()
        if kgid and row_kgid == kgid and not same_record(row_kgid):
            issues.append({"field": "kgid", "message": f"KGID already exists: {kgid} - {row_name}"})
        for field, mobile_value in [("mobile", mobile), ("mobile2", mobile2)]:
            if mobile_value and row_mobile == mobile_value and not same_record(row_kgid) and not any(issue["field"] == field for issue in issues):
                issues.append({"field": field, "message": f"Mobile already exists: {mobile_value} - {row_name or row_kgid}"})
        if any(issue["field"] == "kgid" for issue in issues) and all(any(issue["field"] == f for issue in issues) or not v for f, v in [("mobile", mobile), ("mobile2", mobile2)]):
            break

    if imported is not None:
        imported_headers = headers_for(imported)
        imported_cols = {header: idx + 1 for idx, header in enumerate(imported_headers)}
        for row in range(2, imported.max_row + 1):
            row_kgid = as_text(imported.cell(row, imported_cols.get("kgid", 1)).value).strip()
            row_name = as_text(imported.cell(row, imported_cols.get("name", 2)).value).strip()
            row_mobile1 = as_text(imported.cell(row, imported_cols.get("mobile1", 3)).value).strip()
            row_mobile2 = as_text(imported.cell(row, imported_cols.get("mobile2", 4)).value).strip()
            row_email = as_text(imported.cell(row, imported_cols.get("email", 10)).value).strip().lower()
            if same_record(row_kgid):
                continue
            for field, mobile_value in [("mobile", mobile), ("mobile2", mobile2)]:
                if mobile_value and mobile_value in {row_mobile1, row_mobile2} and not any(issue["field"] == field for issue in issues):
                    issues.append({"field": field, "message": f"Mobile already exists: {mobile_value} - {row_name or row_kgid}"})
            if email and row_email == email and not any(issue["field"] == "email" for issue in issues):
                issues.append({"field": "email", "message": f"Email already exists: {email} - {row_name or row_kgid}"})
            requested_fields = [("mobile", mobile), ("mobile2", mobile2), ("email", email)]
            if all(any(issue["field"] == f for issue in issues) or not v for f, v in requested_fields):
                break

    return {"ok": True, "duplicate": bool(issues), "issues": issues}


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/api/data":
            try:
                payload = load_data()
                body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                body = json.dumps({"error": str(exc)}).encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        if path == "/api/check-duplicate":
            try:
                payload = duplicate_status(parse_qs(parsed.query))
                body = json.dumps(payload).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                body = json.dumps({"ok": False, "error": str(exc)}).encode("utf-8")
                self.send_response(400)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return

        self.serve_file(path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/employee":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                emp_data = payload.get("employee") if isinstance(payload, dict) and "employee" in payload else payload
                result = save_employee(emp_data)
                body = json.dumps(result).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                body = json.dumps({"ok": False, "error": str(exc)}).encode("utf-8")
                self.send_response(400)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        elif parsed.path == "/api/transfer-request":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                tr_data = payload.get("transferRequest") if isinstance(payload, dict) and "transferRequest" in payload else payload
                result = save_transfer_request(tr_data)
                body = json.dumps(result).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                body = json.dumps({"ok": False, "error": str(exc)}).encode("utf-8")
                self.send_response(400)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        self.send_error(404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/employee":
            try:
                query = parse_qs(parsed.query)
                kgid = as_text(query.get("kgid", [""])[0]).strip()
                result = delete_employee(kgid)
                body = json.dumps(result).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                body = json.dumps({"ok": False, "error": str(exc)}).encode("utf-8")
                self.send_response(400)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        elif parsed.path == "/api/transfer-request":
            try:
                query = parse_qs(parsed.query)
                kgid = as_text(query.get("kgid", [""])[0]).strip()
                result = delete_transfer_request(kgid)
                body = json.dumps(result).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                body = json.dumps({"ok": False, "error": str(exc)}).encode("utf-8")
                self.send_response(400)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        self.send_error(404)

    def serve_file(self, path):
        if path in ("/", "/dashboard"):
            file_path = APP_DIR / "index.html"
        else:
            file_path = APP_DIR / path.lstrip("/")
        if not file_path.exists() or not file_path.is_file():
            self.send_error(404)
            return
        body = file_path.read_bytes()
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        return


if __name__ == "__main__":
    # ── Startup backup ───────────────────────────────────────────────
    startup_bk = backup_workbook("startup")
    if startup_bk:
        print(f"[DPDMS] Startup backup created: {startup_bk.name}")
    else:
        print("[DPDMS] WARNING: Data workbook not found – no startup backup created.")
    # ── Launch server ────────────────────────────────────────────────
    server = ThreadingHTTPServer(("127.0.0.1", 8765), Handler)
    print("[DPDMS] Dashboard running at http://127.0.0.1:8765")
    server.serve_forever()
